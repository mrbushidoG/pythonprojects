from openpyxl import Workbook
from string import ascii_uppercase

wb = Workbook()
# ws1 = wb.create_sheet("Range numbers",0)
# for row in range(1,40):
#     ws1.append(range(600))
#
# ws2 = wb.create_sheet("Second Range",1)
#
# for row in range(1,40,2):
#     ws2.append(range(200))
#
# wb.save("C:\\Users\\mrbushido90\\Numbers Range.xlsx")

sheet = wb.active
sheet = wb.create_sheet("Show Letters",0)
for row in range(1,43):
    for c in ascii_uppercase:
        sheet[ c +str(row)] = row
wb.save("C:\\Users\\mrbushido90\\Writing to a cell.xlsx")
wb.close()